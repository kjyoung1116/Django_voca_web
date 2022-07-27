from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from openpyxl.utils.cell import column_index_from_string
from time import sleep

options = webdriver.ChromeOptions()
options.add_argument('--incognito')
'''options.add_argument('--headless')'''
options.add_argument('--no-sandbox')
options.add_argument('--disable-setuid-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_experimental_option("excludeSwitches", ["enable-logging"])

synonyms = []

# 브라우저 생성
browser = webdriver.Chrome('C:/chromedriver.exe', options=options)

# 사이트 열기
browser.get('https://www.onelook.com/thesaurus/')

# 로딩 10초 기다리기
browser.implicitly_wait(10)

# 파일 불러오기

vocas = [] # 단어 모음

wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb['Sheet1']

voca = ws['B'][5000:] # B열만 사용 # 단어 번호 +1

for cell in voca:
    vocas.append(cell.value)



for i in vocas:
    try:
        # 검색창 클릭
        browser.implicitly_wait(1)
        sleep(0.1)
        meaning_search = browser.find_element(By.CSS_SELECTOR, '#thesinput')
        browser.implicitly_wait(1)
        meaning_search.clear()
        browser.implicitly_wait(1)
        meaning_search.click()
        sleep(0.1)
        meaning_search = browser.find_element(By.CSS_SELECTOR, '#thesinput')
        browser.implicitly_wait(1)
        meaning_search.clear()
        browser.implicitly_wait(1)
        meaning_search.click()
        browser.implicitly_wait(1)
        # 검색어 입력
        meaning_search.send_keys(i)
        browser.implicitly_wait(1)
        meaning_search.send_keys(Keys.ENTER)
        browser.implicitly_wait(1)
        # 크롤링할 부분 선택
        sleep(0.1)
        synonym = browser.find_element(By.CSS_SELECTOR, '#zone1 > div > table > tbody > tr > td:nth-child(1)').text

        sleep(0.1)
        browser.implicitly_wait(1)

        synonym = browser.find_element(By.CSS_SELECTOR, '#zone1 > div > table > tbody > tr > td:nth-child(1)').text
        limit_index = synonym.find('6')
        synonym = synonym[:limit_index]
        browser.implicitly_wait(1)
        # synonyms에 의미 저장
        sleep(0.1)
        synonyms.append(synonym)
        browser.implicitly_wait(1)

    except:
        synonyms.append('None')

browser.quit()


# 엑셀에 입력

num2=5000 # 상황에 맞게 바꾸기
for i in synonyms:
    ws.cell(row=num2, column=8, value=i)
    num2+=1


wb.save('voca.xlsx')
wb.close()



