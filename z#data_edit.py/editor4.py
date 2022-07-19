import time
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl.utils.cell import column_index_from_string

# 동의어 보감 중 추가 안 된 단어 추가

wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb['동의어 보감']
voca = ws['C'][1:] # B열만 사용 # 단어 번호 +1

# sheet1
ws2 = wb['Sheet1']

voca_sheet = [cell.value for cell in list(ws2.columns)[1]]

# 추가 어휘 sheet
ws3 = wb['추가 어휘']


vocas = [] # 단어 모음
num = 70
# 단어 있는지 확인 후 없으면 추가
for cell in voca:
    vocas.append(cell.value)
    
for i in vocas:
    voca_list = i.split(', ')
    
    for j in voca_list:

        if j not in voca_sheet:
            ws3.cell(row=num, column=1, value=j)
            num+=1


wb.save('voca.xlsx')
wb.close()
