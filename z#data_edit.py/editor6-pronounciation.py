from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from openpyxl.utils.cell import column_index_from_string
from time import sleep

options = webdriver.ChromeOptions()
options.add_argument('--incognito')
'''options.add_argument('--headless')'''
options.add_argument('--no-sandbox')
options.add_argument('--disable-setuid-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_experimental_option("excludeSwitches", ["enable-logging"])

pronunciations = []

# 브라우저 생성
browser = webdriver.Chrome('C:/chromedriver.exe', options=options)

# 사이트 열기
browser.get('https://dict.naver.com/enendict/#/search?query=outgrowth')

# 로딩 10초 기다리기
browser.implicitly_wait(10)

# 파일 불러오기

vocas = [] # 단어 모음

wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb['Sheet1']

voca = ws['B'][6000:] # B열만 사용 # 단어 번호 +1

for cell in voca:
    vocas.append(cell.value)



for i in vocas:
    try:
        # 검색창 클릭
        browser.implicitly_wait(1)
        sleep(0.08)
        meaning_search = browser.find_element(By.CSS_SELECTOR, '#ac_input')
        browser.implicitly_wait(1)
        meaning_search.clear()
        browser.implicitly_wait(1)
        meaning_search.click()
        sleep(0.02)
        meaning_search = browser.find_element(By.CSS_SELECTOR, '#ac_input')
        browser.implicitly_wait(1)
        meaning_search.clear()
        browser.implicitly_wait(1)
        meaning_search.click()
        browser.implicitly_wait(1)
        # 검색어 입력
        meaning_search.send_keys(i)
        browser.implicitly_wait(1)
        meaning_search.send_keys(Keys.ENTER)
        browser.implicitly_wait(1)
        # 크롤링할 부분 선택
        sleep(0.05)
        pronunciation = browser.find_element(By.CSS_SELECTOR, '#searchPage_entry > div > div:nth-child(1) > div.listen_global_area._globalSearchPronModule.my_global_pron_area > div > div > span.pronounce').text

        browser.implicitly_wait(1)
        pronunciation = browser.find_element(By.CSS_SELECTOR, '#searchPage_entry > div > div:nth-child(1) > div.listen_global_area._globalSearchPronModule.my_global_pron_area > div > div > span.pronounce').text

        browser.implicitly_wait(1)
        # pronunciations에 의미 저장
        sleep(0.05)
        pronunciations.append(pronunciation)
        browser.implicitly_wait(1)

    except:
        pronunciations.append('None')

browser.quit()


# 엑셀에 입력

num2=6001 # 상황에 맞게 바꾸기
for i in pronunciations:
    ws.cell(row=num2, column=9, value=i)
    num2+=1


wb.save('voca.xlsx')
wb.close()
