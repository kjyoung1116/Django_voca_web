from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from openpyxl.utils.cell import column_index_from_string


options = webdriver.ChromeOptions()
options.add_argument('--incognito')
'''options.add_argument('--headless')'''
options.add_argument('--no-sandbox')
options.add_argument('--disable-setuid-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_experimental_option("excludeSwitches", ["enable-logging"])

meanings = []


# 브라우저 생성
browser = webdriver.Chrome('C:/chromedriver.exe', options=options)

# 사이트 열기
browser.get('https://en.dict.naver.com/#/main')

# 로딩 10초 기다리기
browser.implicitly_wait(10)

# 파일 불러오기

vocas = [] # 단어 모음

wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb['추가 어휘']

voca = ws['B'][1:] # B열만 사용 # 단어 번호 +1
mean = ws['c'][1:]

for cell,mean_cell in zip(voca,mean):
    if cell.value != None and cell.value.encode().isalpha() and mean_cell.value == None:
        vocas.append(cell.value)


'''button_close = browser.find_element(By.CSS_SELECTOR, '#platformEventLayerClose')
button_close.click()'''
for i in vocas:
    try:
        if '-' in i or ' ' in i:
            meanings.append('None')
            continue
        
        else:
            # 검색창 클릭
            meaning_search = browser.find_element(By.CSS_SELECTOR, 'input.keyword')

            meaning_search.clear()

            meaning_search.click()
            browser.implicitly_wait(2)
            # 검색어 입력
            meaning_search.send_keys(i)
            browser.implicitly_wait(0.5)
            meaning_search.send_keys(Keys.ENTER)
            browser.implicitly_wait(2)

            # 크롤링하고 싶은 텍스트 선택
            items = browser.find_element(By.CSS_SELECTOR,'.mean_list:not(.mark)').text
            browser.implicitly_wait(2)
            items = items.replace('명사','(n)')
            items = items.replace('대명사','(pron)')
            items = items.replace('동사','(v)')
            items = items.replace('형용사','(adj)')
            items = items.replace('부사','(ad)')
            items = items.replace('전치사','(prep)')
            items = items.replace('접속사','(conj)')
            items = items.replace('감탄사','(exclam)')
            items = items.replace('\n','')
            items = items.replace('1.','')
            items = items.replace('2.','\n')
            items = items.replace('3.','\n')
            items = items.replace('비격식', '[비격식]')
            items = items.replace('격식','[격식]')
            items = items.replace('은어', '[은어]')
            items = items.replace('자(v)','(v)')
            items = items.replace('타(v)','(v)')
            browser.implicitly_wait(2)
            items = items.replace('보통 못마땅함', '')
            items = items.replace('못마땅함', '')
            items = items.replace('전문 용어', '')
            items = items.replace('문예체','')
            items = items.replace('[UC]','')
            items = items.replace('[U]','')
            items = items.replace('호감','')
            items = items.replace('구식','')
            items = items.replace('美','')
            items = items.replace('특히','')
            browser.implicitly_wait(0.5)
            items = items.replace('또는','')
            items = items.replace('英','')
            items = items.replace('속어','')
            items = items.replace('흔히','')
            items = items.replace('옛글투','')
            items = items.replace('英','')
            items = items.replace('[비[격식]]','[비격식]')
            items = items.replace('  ',' ')
            items = items.replace('스코·북잉글','')
            items = items.replace('호주 영어','')
            items = items.replace('두운 :','')
            items = items.replace('( ','(')
            items = items.replace('  ',' ')
            browser.implicitly_wait(2)
            # meanings에 의미 저장
            meanings.append(items)

    except:
        meanings.append('None')
        continue


browser.quit()


# 엑셀에 입력

num=2  # 상황에 맞게 바꾸기 (단어 번호 +1)
for i in meanings:
    ws.cell(row=num, column=3, value=i)
    num+=1

wb.save('voca.xlsx')
wb.close()



