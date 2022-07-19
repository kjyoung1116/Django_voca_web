from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from openpyxl.utils.cell import column_index_from_string



options = webdriver.ChromeOptions()
options.add_argument('--incognito')
'''options.add_argument('--headless')'''
options.add_argument('--no-sandbox')
options.add_argument('--disable-setuid-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_experimental_option("excludeSwitches", ["enable-logging"])

meanings = []

synonyms = []


# 브라우저 생성
browser = webdriver.Chrome('C:/chromedriver.exe', options=options)

# 사이트 열기
browser.get('https://en.dict.naver.com/#/main')

# 로딩 10초 기다리기
browser.implicitly_wait(10)

# 파일 불러오기
excelFile = openpyxl.load_workbook('voca.xlsx')
# 시트 이름으로 불러오기 - sheet = load_wb['Sheet1]


vocas = [] # 단어 모음
pronounces = []
synonyms = []
wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb.active

voca = ws['B'][1:5] # B열만 사용 # 단어 번호 +1

for voca in voca:
        vocas.append(voca.value)


'''button_close = browser.find_element(By.CSS_SELECTOR, '#platformEventLayerClose')
button_close.click()'''
for i in vocas:
    try:
        
        # 검색창 클릭
        meaning_search = browser.find_element(By.CSS_SELECTOR, 'input.keyword')

        meaning_search.clear()

        meaning_search.click()
        browser.implicitly_wait(2)
        # 검색어 입력
        meaning_search.send_keys(i)
        browser.implicitly_wait(0.5)
        meaning_search.send_keys(Keys.ENTER)

        pronounce = browser.find_element(By.CSS_SELECTOR, '#searchPage_entry > div > div:nth-child(1) > div.origin > ul > li:nth-child(1) > span.listen_pronunce_mark').text # 발음기호
       
        # synonyms에 의미 저장
        pronounces.append(pronounce)

    except NoSuchElementException as n:
        synonyms.append('')
        continue

browser.quit()


# 엑셀에 입력

num=2  # 상황에 맞게 바꾸기 (단어 번호 +1)
for i in pronounces:
    ws.cell(row=num, column=9, value=i)
    num+=1


wb.save('voca.xlsx')
wb.close()



