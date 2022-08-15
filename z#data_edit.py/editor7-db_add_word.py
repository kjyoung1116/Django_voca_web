from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from openpyxl.utils.cell import column_index_from_string


options = webdriver.ChromeOptions()
options.add_argument('--incognito')
'''options.add_argument('--headless')'''
options.add_argument('--no-sandbox')
options.add_argument('--disable-setuid-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_experimental_option("excludeSwitches", ["enable-logging"])

wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb['Sheet1']

number = ws['A'][0:10]
voca = ws['B'][0:10]
meaning = ws['C'][0:10]
genre = ws['G'][0:10]
synonyms = ws['H'][0:10]
pronunciations = ws['I'][0:10]

# 브라우저 생성
browser = webdriver.Chrome('C:/chromedriver.exe', options=options)

# 사이트 열기
browser.get('http://127.0.0.1:8000/reviewer/new_card')

# 로딩 10초 기다리기
browser.implicitly_wait(10)


for i,j,k,l,m,n in zip(number, voca,meaning, genre, synonyms,pronunciations):

    number = browser.find_element(By.CSS_SELECTOR, '#id_addition1')
    number.click()
    number.send_keys(i)
    browser.implicitly_wait(10)

    voca = browser.find_element(By.CSS_SELECTOR, '#id_question')
    voca.click()
    voca.send_keys(j)
    browser.implicitly_wait(10)

    meaning = browser.find_element(By.CSS_SELECTOR, '#id_answer')
    meaning.click()
    meaning.send_keys(k)
    browser.implicitly_wait(10)

    genre = browser.find_element(By.CSS_SELECTOR, '#id_addition2')
    genre.click()
    genre.send_keys(l)
    browser.implicitly_wait(10)

    synonyms = browser.find_element(By.CSS_SELECTOR, '#id_addition3')
    synonyms.click()
    synonyms.send_keys(m)
    browser.implicitly_wait(10)

    pronunciations = browser.find_element(By.CSS_SELECTOR, '#id_addition4')
    pronunciations.click()
    pronunciations.send_keys(n)
    browser.implicitly_wait(10)

    pronunciations.send_keys(Keys.ENTER)


