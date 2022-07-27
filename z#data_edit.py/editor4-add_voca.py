import time
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl.utils.cell import column_index_from_string

# 동의어 보감 중 추가 안 된 단어 추가

wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb['동의어 보감']
voca = ws['C'][1:] # B열만 사용 # 단어 번호 +1

# sheet1
ws2 = wb['Sheet1']
ws3 = wb['추가 어휘']

voca_sheet = [cell.value for cell in list(ws2.columns)[1]]
voca_sheet2 = [cell.value for cell in list(ws3.columns)[1]]
# 추가 어휘 sheet



vocas = [] # 단어 모음
num = 776
# 단어 있는지 확인 후 없으면 추가
for cell in voca:
    vocas.append(cell.value)
    
for i in vocas:
    voca_list = i.split(', ')
    
    for j in voca_list:

        if j not in voca_sheet and j not in voca_sheet2:
            ws3.cell(row=num, column=2, value=j)
            num+=1


wb.save('voca.xlsx')
wb.close()


# 숙어 추가
wb = openpyxl.load_workbook('voca.xlsx') 
ws = wb['동의어 보감']
idiom = ws['d'][1:] # B열만 사용 # 단어 번호 +1


ws2 = wb['숙어']

idiom_sheet = [cell.value for cell in list(ws2.columns)[0]]

idioms = [] # 숙어 모음
num = 60
# 단어 있는지 확인 후 없으면 추가
for cell in idiom:
    idioms.append(cell.value)
    
for i in idioms:
    if i.find(',') == 0:
        idiom_list = i
    else:
        idiom_list = i.split(', ')
    
    for j in idiom_list:

        if j not in idiom_sheet:
            ws2.cell(row=num, column=7, value=j)
            num+=1

wb.save('voca.xlsx')
wb.close()
