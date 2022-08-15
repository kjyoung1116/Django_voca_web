from audioop import add
from msilib import CreateRecord
from pydoc import plain
from re import M
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.models import User as user_model
from soupsieve import select
from reviewer.forms import CreateReviewSettingsForm
from .models import review_settings, Card
import pandas as pd
from django.views.generic import (ListView,CreateView,UpdateView,DetailView)
from django.urls import reverse_lazy
from .forms import CardCheckForm
import openpyxl
import random
from datetime import datetime, timedelta

# Create your views here.
def index(request):
    user = get_object_or_404(user_model, pk=request.user.id)
    my_reviewers = review_settings.objects.filter(user_id = user)
    return render(request, '../templates/reviewer.html', {'my_reviewers': my_reviewers})

def reviewer_create(request):
    
    if request.method == 'GET':
        form = CreateReviewSettingsForm()
        return render(request, '../templates/reviewer_create.html', {'form': form})

    elif request.method == 'POST':
        user = get_object_or_404(user_model, pk=request.user.id)
        
        plan_name = request.POST['plan_name']
        selected_voca = request.POST['selected_voca']
        voca_per_day = request.POST['voca_per_day']
        reviewing_intensity = request.POST['reviewing_intensity']
        purpose = request.POST['purpose']

        new_setting = review_settings.objects.create(
            user = user,
            plan_name = plan_name,
            selected_voca = selected_voca,
            voca_per_day = voca_per_day,
            reviewing_intensity = reviewing_intensity,
            purpose = purpose
        )

        new_setting.save()

        wb = openpyxl.load_workbook('voca.xlsx') 
        ws = wb['Sheet1']
# 데이터셋마다 변경
        number = ws['A'][1:7590]
        voca = ws['B'][1:7590]
        meaning = ws['C'][1:7590]
        genre = ws['G'][1:7590]
        synonyms = ws['H'][1:7590]
        pronunciations = ws['I'][1:7590]

### 모델에 복습 날짜 추가, 복습 날짜가 오늘인 단어만 출력, knowing_level에 따라 복습 날짜 재설정. box 2,3,4,5는 복습 날짜가 각각 n일 후인 단어들 담고 있음.

        if selected_voca == 'all':
            for i,j,k,l,m,n in zip(voca,meaning,number,genre,synonyms,pronunciations):
                new_card = Card.objects.create(
                    question = i.value,
                    answer = j.value,
                    addition0 = user,
                    addition1 = k.value,
                    addition2 = l.value,
                    addition3 = m.value,
                    addition4 = n.value,
                    addition9 = plan_name
                )
            return redirect('reviewer:reviewer_current')
        return render(request, '../templates/reviewer_plan_detail.html')


def my_reviewers(request):
    user = get_object_or_404(user_model, pk=request.user.id)
    my_reviewers = review_settings.objects.filter(user_id = user)

    return render(request, '../templates/reviewer_current.html', {'my_reviewers': my_reviewers})


'''def reviewer_plan_detail(request, plan_name):
    user = get_object_or_404(user_model, pk=request.user.id)
    card = Card.objects.filter(addition9 = plan_name, addition0 = user)[0:1]

    return render(request, '../templates/reviewer_plan_detail.html', {'card': card})'''

class reviewer_plan_detail(ListView):

    model = Card
    template_name = 'reviewer_plan_detail.html'

    def get_context_data(self, **kwargs):
        user = self.request.user
        context = super(reviewer_plan_detail, self).get_context_data(**kwargs)
        context['card'] = Card.objects.filter(addition0 = user)
        dates = set(Card.objects.filter(addition0 = user, addition9 = self.kwargs['plan_name']).values_list('review_date', flat=True))
        context["review_date"] = sorted(dates)
        return context
###### 플랜별/유저별로 구분하는 것 필요.
class reviewer_detail(ListView):
    model = Card, review_settings
    queryset = Card.objects.all().order_by("box", "-date_created")[0:0]
    template_name = 'reviewer_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super(reviewer_detail, self).get_context_data(**kwargs)
        context['card_with_plan'] = Card.objects.filter(addition0 = self.request.user)[0:1]
        
        return context

class new_card(CreateView):
    model = Card
    fields = ["question", "answer", "box", "addition1", "addition2", "addition3", "addition4", "addition5", "addition6", "addition7", "addition8", "addition9", "addition0"]
    success_url = reverse_lazy("reviewer:new_card")
    template_name = 'reviewer_new_card.html'
    

class update_card(new_card, UpdateView):
    success_url = reverse_lazy("reviewer:reviewer_detail")
    template_name = 'reviewer_update_card.html'


class box_view(reviewer_detail):
    template_name = "reviewer_box.html"
    form_class = CardCheckForm
    
    def get_queryset(self):
        user = self.request.user
        return Card.objects.filter(review_date = self.kwargs['review_date'], addition0 = user)[0:1]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        context = super().get_context_data(**kwargs)
        user = self.request.user

        for i in set(Card.objects.filter(addition0 = user).values_list('review_date', flat=True)):
            context["box_count"] = len(Card.objects.filter(review_date = self.kwargs['review_date'], addition0 = user))

        context["box_name"] = self.kwargs['review_date']+' 단어'

        if self.object_list:
            context["check_card"] = self.object_list

        dates = set(Card.objects.filter(addition0 = user).values_list('review_date', flat=True))
        context["review_date"] = sorted(dates)
        return context


    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            card_id = get_object_or_404(Card, id = form.cleaned_data["card_id"])
            knowing_level = form.cleaned_data["knowing_level"]
            card = Card.objects.get(question = card_id, addition0 = self.request.user)
            card.box = card.box + knowing_level
            card.review_date = card.review_date + timedelta(days=knowing_level)
            card.save()

        return redirect(request.META.get("HTTP_REFERER"))

###### 박스 이름을 복습할 날짜로 바꾸고, 박스가 계속 생성되게 만들기. 빈 박스이며, 오늘보다 이전 날짜인 경우 삭제. 오늘 날짜인 박스 위에 오늘 표시.
